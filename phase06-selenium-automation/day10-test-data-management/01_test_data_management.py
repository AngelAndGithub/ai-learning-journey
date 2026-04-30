# Day 10: 测试数据管理
# 测试数据管理示例

print("Day 10: 测试数据管理")

import os
import json
import csv
from openpyxl import load_workbook
from faker import Faker

# 1. Excel数据驱动
print("\n1. Excel数据驱动:")

class ExcelDataManager:
    """Excel数据管理类"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def read_data(self, sheet_name):
        """读取Excel数据"""
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook[sheet_name]
            
            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # 读取数据
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
            
            return data
        except Exception as e:
            print(f"   - 读取Excel失败: {e}")
            return []
    
    def write_data(self, sheet_name, data):
        """写入Excel数据"""
        try:
            from openpyxl import Workbook
            
            # 创建工作簿
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
            
            if data:
                # 写入表头
                headers = list(data[0].keys())
                for i, header in enumerate(headers, 1):
                    sheet.cell(row=1, column=i, value=header)
                
                # 写入数据
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"   - 写入Excel失败: {e}")
            return False

# 创建测试数据
print("   - 创建测试数据:")
test_data = [
    {"username": "user1", "password": "pass1", "expected": "success"},
    {"username": "user2", "password": "pass2", "expected": "success"},
    {"username": "invalid", "password": "wrong", "expected": "failure"}
]

# 写入Excel
excel_manager = ExcelDataManager("test_data.xlsx")
excel_manager.write_data("LoginData", test_data)
print("   - Excel数据写入成功")

# 读取Excel
read_data = excel_manager.read_data("LoginData")
print("   - 读取Excel数据:")
for row in read_data:
    print(f"   {row}")

# 2. JSON配置文件
print("\n2. JSON配置文件:")

class JSONConfigManager:
    """JSON配置管理类"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def load_config(self):
        """加载配置"""
        try:
            if os.path.exists(self.file_path):
                with open(self.file_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            else:
                return {}
        except Exception as e:
            print(f"   - 加载JSON配置失败: {e}")
            return {}
    
    def save_config(self, config):
        """保存配置"""
        try:
            with open(self.file_path, "w", encoding="utf-8") as f:
                json.dump(config, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"   - 保存JSON配置失败: {e}")
            return False

# 创建配置
config = {
    "environments": {
        "dev": {
            "base_url": "https://dev.example.com",
            "username": "dev_user",
            "password": "dev_pass"
        },
        "test": {
            "base_url": "https://test.example.com",
            "username": "test_user",
            "password": "test_pass"
        },
        "prod": {
            "base_url": "https://prod.example.com",
            "username": "prod_user",
            "password": "prod_pass"
        }
    },
    "browser": "chrome",
    "timeout": 10
}

# 保存配置
json_manager = JSONConfigManager("config.json")
json_manager.save_config(config)
print("   - JSON配置保存成功")

# 加载配置
loaded_config = json_manager.load_config()
print("   - 加载JSON配置:")
print(f"   浏览器: {loaded_config.get('browser')}")
print(f"   超时时间: {loaded_config.get('timeout')}秒")
print("   环境配置:")
for env, env_config in loaded_config.get('environments', {}).items():
    print(f"     {env}: {env_config.get('base_url')}")

# 3. CSV数据文件
print("\n3. CSV数据文件:")

class CSVDataManager:
    """CSV数据管理类"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def read_data(self):
        """读取CSV数据"""
        try:
            data = []
            with open(self.file_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    data.append(row)
            return data
        except Exception as e:
            print(f"   - 读取CSV失败: {e}")
            return []
    
    def write_data(self, data):
        """写入CSV数据"""
        try:
            if data:
                headers = list(data[0].keys())
                with open(self.file_path, "w", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=headers)
                    writer.writeheader()
                    writer.writerows(data)
            return True
        except Exception as e:
            print(f"   - 写入CSV失败: {e}")
            return False

# 写入CSV
csv_manager = CSVDataManager("test_data.csv")
csv_manager.write_data(test_data)
print("   - CSV数据写入成功")

# 读取CSV
csv_data = csv_manager.read_data()
print("   - 读取CSV数据:")
for row in csv_data:
    print(f"   {row}")

# 4. 测试数据生成
print("\n4. 测试数据生成:")

class TestDataGenerator:
    """测试数据生成器"""
    def __init__(self):
        """初始化"""
        self.faker = Faker("zh_CN")
    
    def generate_user_data(self, count=5):
        """生成用户数据"""
        users = []
        for _ in range(count):
            user = {
                "username": self.faker.user_name(),
                "email": self.faker.email(),
                "password": self.faker.password(),
                "name": self.faker.name(),
                "phone": self.faker.phone_number(),
                "address": self.faker.address()
            }
            users.append(user)
        return users
    
    def generate_product_data(self, count=5):
        """生成产品数据"""
        products = []
        for _ in range(count):
            product = {
                "name": self.faker.word(),
                "price": round(self.faker.random_number(digits=3), 2),
                "description": self.faker.sentence(),
                "category": self.faker.word(),
                "stock": self.faker.random_number(digits=2)
            }
            products.append(product)
        return products

# 生成测试数据
generator = TestDataGenerator()
user_data = generator.generate_user_data(3)
print("   - 生成用户数据:")
for user in user_data:
    print(f"   用户名: {user['username']}, 邮箱: {user['email']}")

product_data = generator.generate_product_data(3)
print("   - 生成产品数据:")
for product in product_data:
    print(f"   产品: {product['name']}, 价格: {product['price']}")

# 5. 参数化测试
print("\n5. 参数化测试:")

# 模拟参数化测试框架
def parameterized_test(data):
    """参数化测试"""
    results = []
    for test_case in data:
        print(f"   测试用例: {test_case}")
        # 模拟测试逻辑
        username = test_case.get("username")
        password = test_case.get("password")
        expected = test_case.get("expected")
        
        # 模拟测试结果
        if expected == "success":
            results.append({"case": test_case, "result": "PASS"})
            print(f"   结果: PASS")
        else:
            results.append({"case": test_case, "result": "FAIL"})
            print(f"   结果: FAIL")
    return results

# 从Excel读取数据进行参数化测试
print("   - 从Excel读取数据进行参数化测试:")
excel_data = excel_manager.read_data("LoginData")
parameterized_test(excel_data)

# 从JSON读取数据进行参数化测试
print("   - 从JSON读取数据进行参数化测试:")
json_data = json_manager.load_config().get("test_cases", [])
if not json_data:
    # 如果没有测试用例，创建一些
    json_data = test_data
    config["test_cases"] = json_data
    json_manager.save_config(config)
    parameterized_test(json_data)

# 6. 数据管理最佳实践
print("\n6. 数据管理最佳实践:")
print("   - 数据与代码分离")
print("   - 使用适合的数据格式（Excel、JSON、CSV）")
print("   - 测试数据的可维护性")
print("   - 数据的版本控制")
print("   - 测试数据的安全性（避免硬编码敏感信息）")
print("   - 动态生成测试数据")

# 7. 实际应用示例
print("\n7. 实际应用示例:")

class TestRunner:
    """测试运行器"""
    def __init__(self, data_source):
        """初始化"""
        self.data_source = data_source
    
    def run_tests(self):
        """运行测试"""
        if self.data_source == "excel":
            data = excel_manager.read_data("LoginData")
        elif self.data_source == "json":
            data = json_manager.load_config().get("test_cases", [])
        elif self.data_source == "csv":
            data = csv_manager.read_data()
        else:
            data = test_data
        
        return parameterized_test(data)

# 运行测试
print("   - 从不同数据源运行测试:")
runner = TestRunner("excel")
runner.run_tests()

print("\n测试数据管理示例完成！")
print("\n学习要点：")
print("1. Excel数据驱动:")
print("   - 使用openpyxl库读取和写入Excel文件")
print("   - 支持复杂的测试数据结构")
print("2. JSON配置文件:")
print("   - 适合存储配置信息")
print("   - 支持嵌套结构")
print("3. CSV数据文件:")
print("   - 简单易用")
print("   - 适合简单的测试数据")
print("4. 测试数据生成:")
print("   - 使用Faker库生成随机测试数据")
print("   - 支持多种数据类型")
print("5. 参数化测试:")
print("   - 从外部数据源读取测试数据")
print("   - 提高测试覆盖率")
print("6. 最佳实践:")
print("   - 数据与代码分离")
print("   - 数据的可维护性")
print("   - 数据的安全性")