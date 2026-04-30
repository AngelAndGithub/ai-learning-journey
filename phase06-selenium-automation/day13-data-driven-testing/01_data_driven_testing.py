# Day 13: 数据驱动测试
# 数据驱动测试示例

print("Day 13: 数据驱动测试")

import os
import json
import csv
from openpyxl import load_workbook
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# 1. 数据驱动测试概述
print("\n1. 数据驱动测试概述:")
print("   - 数据驱动测试是一种测试方法，其中测试数据与测试逻辑分离")
print("   - 测试数据存储在外部数据源中（Excel、CSV、JSON等）")
print("   - 相同的测试逻辑可以使用不同的测试数据运行")
print("   - 优势：提高测试覆盖率、减少代码重复、便于维护")

# 2. 从Excel读取测试数据
print("\n2. 从Excel读取测试数据:")

class ExcelDataReader:
    """Excel数据读取器"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def read_data(self, sheet_name):
        """读取Excel数据"""
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook[sheet_name]
            
            # 读取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # 读取数据
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
            
            return data
        except Exception as e:
            print(f"   - 读取Excel失败: {e}")
            return []

# 创建测试数据文件
print("   - 创建测试数据文件:")

# 写入Excel测试数据
try:
    from openpyxl import Workbook
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "LoginData"
    
    # 写入表头
    headers = ["username", "password", "expected", "test_case"]
    for i, header in enumerate(headers, 1):
        sheet.cell(row=1, column=i, value=header)
    
    # 写入测试数据
    test_data = [
        ["standard_user", "secret_sauce", "success", "正常登录"],
        ["invalid_user", "secret_sauce", "failure", "无效用户名"],
        ["standard_user", "wrong_password", "failure", "无效密码"],
        ["", "", "failure", "空用户名密码"]
    ]
    
    for row_idx, row in enumerate(test_data, 2):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    workbook.save("login_test_data.xlsx")
    print("   ✓ Excel测试数据文件创建成功")
except Exception as e:
    print(f"   ✗ 创建Excel文件失败: {e}")

# 读取Excel测试数据
excel_reader = ExcelDataReader("login_test_data.xlsx")
excel_data = excel_reader.read_data("LoginData")
print("   - 读取Excel测试数据:")
for row in excel_data:
    print(f"   {row}")

# 3. 从CSV读取测试数据
print("\n3. 从CSV读取测试数据:")

class CSVDataReader:
    """CSV数据读取器"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def read_data(self):
        """读取CSV数据"""
        try:
            data = []
            with open(self.file_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    data.append(row)
            return data
        except Exception as e:
            print(f"   - 读取CSV失败: {e}")
            return []

# 写入CSV测试数据
try:
    with open("login_test_data.csv", "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["username", "password", "expected", "test_case"])
        writer.writeheader()
        writer.writerows([
            {"username": "standard_user", "password": "secret_sauce", "expected": "success", "test_case": "正常登录"},
            {"username": "invalid_user", "password": "secret_sauce", "expected": "failure", "test_case": "无效用户名"},
            {"username": "standard_user", "password": "wrong_password", "expected": "failure", "test_case": "无效密码"},
            {"username": "", "password": "", "expected": "failure", "test_case": "空用户名密码"}
        ])
    print("   ✓ CSV测试数据文件创建成功")
except Exception as e:
    print(f"   ✗ 创建CSV文件失败: {e}")

# 读取CSV测试数据
csv_reader = CSVDataReader("login_test_data.csv")
csv_data = csv_reader.read_data()
print("   - 读取CSV测试数据:")
for row in csv_data:
    print(f"   {row}")

# 4. 从JSON读取测试数据
print("\n4. 从JSON读取测试数据:")

class JSONDataReader:
    """JSON数据读取器"""
    def __init__(self, file_path):
        """初始化"""
        self.file_path = file_path
    
    def read_data(self):
        """读取JSON数据"""
        try:
            with open(self.file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"   - 读取JSON失败: {e}")
            return []

# 写入JSON测试数据
try:
    test_data = [
        {"username": "standard_user", "password": "secret_sauce", "expected": "success", "test_case": "正常登录"},
        {"username": "invalid_user", "password": "secret_sauce", "expected": "failure", "test_case": "无效用户名"},
        {"username": "standard_user", "password": "wrong_password", "expected": "failure", "test_case": "无效密码"},
        {"username": "", "password": "", "expected": "failure", "test_case": "空用户名密码"}
    ]
    
    with open("login_test_data.json", "w", encoding="utf-8") as f:
        json.dump(test_data, f, ensure_ascii=False, indent=2)
    print("   ✓ JSON测试数据文件创建成功")
except Exception as e:
    print(f"   ✗ 创建JSON文件失败: {e}")

# 读取JSON测试数据
json_reader = JSONDataReader("login_test_data.json")
json_data = json_reader.read_data()
print("   - 读取JSON测试数据:")
for row in json_data:
    print(f"   {row}")

# 5. 使用pytest参数化实现数据驱动测试
print("\n5. 使用pytest参数化实现数据驱动测试:")

# 从Excel读取数据进行参数化测试
print("   - 从Excel读取数据进行参数化测试:")

excel_test_data = excel_reader.read_data("LoginData")

@pytest.mark.parametrize("test_data", excel_test_data)
def test_login_from_excel(test_data):
    """从Excel读取数据进行登录测试"""
    print(f"   测试用例: {test_data['test_case']}")
    print(f"   用户名: {test_data['username']}, 密码: {test_data['password']}, 预期: {test_data['expected']}")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.implicitly_wait(10)
    
    try:
        # 打开登录页面
        driver.get("https://www.saucedemo.com/")
        
        # 输入用户名和密码
        driver.find_element(By.ID, "user-name").send_keys(test_data['username'])
        driver.find_element(By.ID, "password").send_keys(test_data['password'])
        driver.find_element(By.ID, "login-button").click()
        
        if test_data['expected'] == "success":
            # 验证登录成功
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "inventory_list"))
            )
            assert "inventory.html" in driver.current_url
            print("   ✓ 测试通过")
        else:
            # 验证登录失败
            error_message = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "error-message-container"))
            )
            assert error_message.is_displayed()
            print("   ✓ 测试通过")
    finally:
        driver.quit()

# 6. 数据驱动测试的最佳实践
print("\n6. 数据驱动测试的最佳实践:")
print("   - 测试数据与测试逻辑分离")
print("   - 使用合适的数据源格式（Excel适合复杂数据，JSON适合结构化数据，CSV适合简单数据）")
print("   - 为测试数据添加描述性字段（如test_case名称）")
print("   - 对测试数据进行版本控制")
print("   - 测试数据应该覆盖各种场景（正常、异常、边界值）")
print("   - 数据读取逻辑应该封装成可重用的函数")

# 7. 多环境测试数据管理
print("\n7. 多环境测试数据管理:")

# 环境配置
environments = {
    "dev": {
        "base_url": "https://dev.saucedemo.com",
        "username": "dev_user",
        "password": "dev_pass"
    },
    "test": {
        "base_url": "https://test.saucedemo.com",
        "username": "test_user",
        "password": "test_pass"
    },
    "prod": {
        "base_url": "https://www.saucedemo.com",
        "username": "standard_user",
        "password": "secret_sauce"
    }
}

# 保存环境配置
try:
    with open("environments.json", "w", encoding="utf-8") as f:
        json.dump(environments, f, ensure_ascii=False, indent=2)
    print("   ✓ 环境配置文件创建成功")
except Exception as e:
    print(f"   ✗ 创建环境配置文件失败: {e}")

# 读取环境配置
with open("environments.json", "r", encoding="utf-8") as f:
    loaded_environments = json.load(f)

print("   - 环境配置:")
for env, config in loaded_environments.items():
    print(f"     {env}: {config['base_url']}")

# 8. 数据驱动测试框架
print("\n8. 数据驱动测试框架:")

class DataDrivenTestFramework:
    """数据驱动测试框架"""
    
    def __init__(self, data_source, data_format):
        """初始化"""
        self.data_source = data_source
        self.data_format = data_format
        self.test_data = []
    
    def load_data(self):
        """加载测试数据"""
        if self.data_format == "excel":
            reader = ExcelDataReader(self.data_source)
            self.test_data = reader.read_data("LoginData")
        elif self.data_format == "csv":
            reader = CSVDataReader(self.data_source)
            self.test_data = reader.read_data()
        elif self.data_format == "json":
            reader = JSONDataReader(self.data_source)
            self.test_data = reader.read_data()
        else:
            print("   - 不支持的数据格式")
    
    def run_tests(self):
        """运行测试"""
        for test_data in self.test_data:
            print(f"   运行测试: {test_data.get('test_case', '未命名测试')}")
            # 这里可以调用测试函数
            self.run_single_test(test_data)
    
    def run_single_test(self, test_data):
        """运行单个测试"""
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        driver.implicitly_wait(10)
        
        try:
            # 打开登录页面
            driver.get("https://www.saucedemo.com/")
            
            # 输入用户名和密码
            driver.find_element(By.ID, "user-name").send_keys(test_data['username'])
            driver.find_element(By.ID, "password").send_keys(test_data['password'])
            driver.find_element(By.ID, "login-button").click()
            
            if test_data['expected'] == "success":
                # 验证登录成功
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "inventory_list"))
                )
                assert "inventory.html" in driver.current_url
                print("   ✓ 测试通过")
            else:
                # 验证登录失败
                error_message = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "error-message-container"))
                )
                assert error_message.is_displayed()
                print("   ✓ 测试通过")
        except Exception as e:
            print(f"   ✗ 测试失败: {e}")
        finally:
            driver.quit()

# 测试数据驱动测试框架
print("   - 测试数据驱动测试框架:")

# 从Excel加载数据
framework = DataDrivenTestFramework("login_test_data.xlsx", "excel")
framework.load_data()
print(f"   从Excel加载了 {len(framework.test_data)} 条测试数据")

# 从CSV加载数据
framework = DataDrivenTestFramework("login_test_data.csv", "csv")
framework.load_data()
print(f"   从CSV加载了 {len(framework.test_data)} 条测试数据")

# 从JSON加载数据
framework = DataDrivenTestFramework("login_test_data.json", "json")
framework.load_data()
print(f"   从JSON加载了 {len(framework.test_data)} 条测试数据")

# 运行测试
print("   - 运行测试:")
framework.run_tests()

# 9. 数据驱动测试的高级应用
print("\n9. 数据驱动测试的高级应用:")
print("   - 动态生成测试数据")
print("   - 测试数据的依赖关系管理")
print("   - 测试数据的验证和清理")
print("   - 与CI/CD集成")

# 10. 实际应用示例
print("\n10. 实际应用示例:")

# 电商网站测试数据
ecommerce_test_data = [
    {"product": "Sauce Labs Backpack", "quantity": 1, "expected_price": "$29.99"},
    {"product": "Sauce Labs Bolt T-Shirt", "quantity": 2, "expected_price": "$15.99"},
    {"product": "Sauce Labs Onesie", "quantity": 3, "expected_price": "$7.99"}
]

# 保存电商测试数据
try:
    with open("ecommerce_test_data.json", "w", encoding="utf-8") as f:
        json.dump(ecommerce_test_data, f, ensure_ascii=False, indent=2)
    print("   ✓ 电商测试数据文件创建成功")
except Exception as e:
    print(f"   ✗ 创建电商测试数据文件失败: {e}")

print("\n数据驱动测试示例完成！")
print("\n学习要点：")
print("1. 数据驱动测试的概念:")
print("   - 测试数据与测试逻辑分离
   - 提高测试覆盖率
   - 减少代码重复
2. 数据源类型:")
print("   - Excel: 适合复杂数据结构
   - CSV: 适合简单数据
   - JSON: 适合结构化数据
3. 实现方法:")
print("   - 使用pytest的参数化装饰器
   - 自定义数据读取器
   - 构建数据驱动测试框架
4. 最佳实践:")
print("   - 测试数据应该覆盖各种场景
   - 为测试数据添加描述性信息
   - 对测试数据进行版本控制
   - 数据读取逻辑封装成可重用函数")